import openpyxl
import datetime
import tabulate

# Проверка перед началом фильтра:
print("ВНИМАНИЕ! Перед обработкой таблица должна соответствовать следующим параметрам:\n\n"
      "************ УКАЗАННЫЕ НИЖЕ СТРАНИЦЫ ДОЛЖНЫ БЫТЬ В ТАБЛИЦЕ В ТАКОМ ПОРЯДКЕ *************")

print(tabulate.tabulate([
        ["Адреса", "Брянск", "АБДЦ",  "АП СООП", "АП ГИБДД", "ЗАПРЕТНИКИ", "ЗАДЕРЖАНИЯ", "ЗАГС"],
],
        tablefmt="simple_outline",
        maxcolwidths=[None, 25])
    )

print("\n********************************** ОФОРМЛЕНИЕ СТРАНИЦ **********************************")

print(tabulate.tabulate([
        ["БРЯНСК",     "+",  "",   "Фамилия", "Имя", "Отчество", "формат: ДД.ММ.ГГГГ"],
        ["АБДЦ",       "+",  "",   "Фамилия", "Имя", "Отчество", "формат: ГГГГММДД"],
        ["АП СООП",    "",   "",   "Фамилия", "Имя", "Отчество", "формат: ДД.ММ.ГГГГ"],
        ["АП ГИБДД",   "+",  "+",  "Фамилия", "Имя", "Отчество", "формат: ДД.ММ.ГГГГ"],
        ["ЗАПРЕТНИКИ", "+",  "",   "Фамилия", "Имя", "Отчество", "формат: ДД.ММ.ГГГГ"],
        ["ЗАДЕРЖАНИЯ", "",   "+",  "Фамилия", "Имя", "Отчество", "формат: ДД.ММ.ГГГГ"],
        ["ЗАГС",       "+",  "+",  "Фамилия", "Имя", "Отчество", "формат: ДД.ММ.ГГГГ"],
],
        headers=["Страница", "Заголовок", "№", "Фамилия", "Имя", "Отчество", "Дата рождения"],
        tablefmt="simple_outline",
        maxcolwidths=[None, 25])
    )

source_file_name = str(input("\nВведите название файла для обработки (только формат xlsx): "))

# Если название файл введено без расширения - оно добавляется
if ".xlsx" in source_file_name:
    pass
else:
    source_file_name = source_file_name + ".xlsx"

# Открываем исходный файл *.xlsx и получаем доступ к страницам
try:
    wb_source            = openpyxl.load_workbook(source_file_name)
    sheet_source_bryansk = wb_source.worksheets[1]
    sheet_source_abdc    = wb_source.worksheets[2]
    sheet_source_soop    = wb_source.worksheets[3]
    sheet_source_gibdd   = wb_source.worksheets[4]
    sheet_source_zapret  = wb_source.worksheets[5]
    sheet_source_zaderj  = wb_source.worksheets[6]
    sheet_source_zags    = wb_source.worksheets[7]
except FileNotFoundError:
    print("Введено неверное название файла или указанный файл отсутсует в текущей дирректории!")
    input("")
    exit()
except IndexError:
    input("Ошибка! В исходном файле не 8 страниц. Проверьте оформление файла!\n"
          "Для выхода введите любой другой символ!")
    exit()

# Создаем итоговый файл xlsx для копирования в него строк из исходного файла после фильтра
wb_result           = openpyxl.Workbook()
sheet_result_abdc   = wb_result.create_sheet("АБДЦ", 0)
sheet_result_soop   = wb_result.create_sheet("АП СООП", 1)
sheet_result_gibdd  = wb_result.create_sheet("АП ГИБДД", 2)
sheet_result_zapret = wb_result.create_sheet("ЗАПРЕТНИКИ", 3)
sheet_result_zaderj = wb_result.create_sheet("ЗАДЕРЖАНИЯ", 4)
sheet_result_zags   = wb_result.create_sheet("ЗАГС", 5)

# Получаем по 1 строке из кажой страницы для предварительного сравнения строк в страницах
test_bryansk = list(sheet_source_bryansk.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True))[0]
test_abdc    = list(sheet_source_abdc.iter_rows(   min_row=2, min_col=1, max_col=4, values_only=True))[0]
test_soop    = list(sheet_source_soop.iter_rows(   min_row=1, min_col=1, max_col=4, values_only=True))[0]
test_gibdd   = list(sheet_source_gibdd.iter_rows(  min_row=2, min_col=2, max_col=5, values_only=True))[0]
test_zapret  = list(sheet_source_zapret.iter_rows( min_row=2, min_col=1, max_col=4, values_only=True))[0]
test_zaderj  = list(sheet_source_zaderj.iter_rows( min_row=1, min_col=2, max_col=5, values_only=True))[0]
test_zags    = list(sheet_source_zags.iter_rows(   min_row=2, min_col=2, max_col=5, values_only=True))[0]

# Создаем множества для удаления одинаковых записей из отфильтрованных списков
set_abdc   = set()
set_soop   = set()
set_gibdd  = set()
set_zapret = set()
set_zaderj = set()
set_zags   = set()

print("\n*************** ПРЕДВАРИТЕЛЬНОЕ СРАВНЕНИЕ СТРОК В СТРАНИЦАХ **************")

try:
    print(tabulate.tabulate([
        ["БРЯНСК", test_bryansk[0], test_bryansk[1], test_bryansk[2], test_bryansk[3].strftime('%d.%m.%Y')],
        ["АБДЦ", test_abdc[0], test_abdc[1], test_abdc[2], str(test_abdc[3])[6:8] + '.' + str(test_abdc[3])[4:6] + '.' + str(test_abdc[3])[:4]],
        ["АП СООП", test_soop[0], test_soop[1], test_soop[2], test_soop[3].strftime('%d.%m.%Y')],
        ["АП ГИБДД", test_gibdd[0], test_gibdd[1], test_gibdd[2], test_gibdd[3]],
        ["ЗАПРЕТНИКИ", test_zapret[0], test_zapret[1], test_zapret[2], test_zapret[3].strftime('%d.%m.%Y')],
        ["ЗАДЕРЖАНИЯ", test_zaderj[0], test_zaderj[1], test_zaderj[2], test_zaderj[3].strftime('%d.%m.%Y')],
        ["ЗАГС", test_zags[0], test_zags[1], test_zags[2], test_zags[3]],
    ],
        headers=["Страница", "Фамилия", "Имя", "Отчество", "Дата рождения"],
        tablefmt="simple_outline",
        maxcolwidths=[None, 25])
    )
except AttributeError:
    input("Исходный файл не соответвсует правилам оформления!\nНажмите Enter для выхода.")
    exit()

question = int(input("\nЕсли таблица и строки в страницах соответствует указанным требованиям - введите цифру 1.\n"
                     "Для выхода из программы нажмите Enter: "))
print("")

if question != 1:
    exit()

# Фильтр страницы АБДЦ
for row_abdc in sheet_source_abdc.iter_rows(min_row=2, values_only=True):
    list_abdc = [cell for cell in row_abdc]
    # Приводим формат даты рождения к ДД.ММ.ГГГГ (только если поле не пустое)
    if list_abdc[3] != None:
        list_abdc[3] = str(list_abdc[3])[6:8] + '.' + str(list_abdc[3])[4:6] + '.' + str(list_abdc[3])[:4]
    #print('АДБЦ:', list_abdc)

    # Сравниваем ФИО и дату рождения между АБДЦ и Брянск
    for row_bryansk in sheet_source_bryansk.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        list_bryansk = [cell for cell in row_bryansk]
        # Приводим формат даты рождения к ДД.ММ.ГГГГ (только если поле не пустое)
        if list_bryansk[3] != None:
            list_bryansk[3] = list_bryansk[3].strftime("%d.%m.%Y")
        #print('Bryansk', list_bryansk)

        # Если строка имеет совпадения по ФИО и дате рождения, то сохраяем эту строку в итоговый файл
        if (list_abdc[0] == list_bryansk[0] and
                list_abdc[1] == list_bryansk[1] and
                list_abdc[2] == list_bryansk[2] and
                list_abdc[3] == list_bryansk[3]):
            set_abdc.add(row_abdc)
# Добавляем уникальные записи в новую страницу
for el in set_abdc:
    print(f"Совпадение АБДЦ с Брянск: {el[0], el[1], el[2], el[3]}")
    sheet_result_abdc.append(el)

# Фильтр страницы АП СООП
for row_soop in sheet_source_soop.iter_rows(min_row=1, values_only=True):
    list_soop = [cell for cell in row_soop]
    # Приводим формат даты рождения к ДД.ММ.ГГГГ (только если поле не пустое)
    if list_soop[3] != None:
        list_soop[3] = list_soop[3].strftime("%d.%m.%Y")
    #print('ОП СООП:', list_soop)

    # Сравниваем ФИО и дату рождения между АБДЦ и Брянск
    for row_bryansk in sheet_source_bryansk.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        list_bryansk = [cell for cell in row_bryansk]
        # Приводим формат даты рождения к ДД.ММ.ГГГГ (только если поле не пустое)
        if list_bryansk[3] != None:
            list_bryansk[3] = list_bryansk[3].strftime("%d.%m.%Y")
        #print('Bryansk', list_bryansk)
        # Если строка имеет совпадения по ФИО и дате рождения, то сохраяем эту строку в итоговый файл
        if (list_soop[0] == list_bryansk[0] and
                list_soop[1] == list_bryansk[1] and
                list_soop[2] == list_bryansk[2] and
                list_soop[3] == list_bryansk[3]):
            set_soop.add(row_soop)
# Добавляем уникальные записи в новую страницу
for el in set_soop:
    print(f"Совпадение АП СООП с Брянск: {el[0], el[1], el[2], el[3]}")
    sheet_result_soop.append(el)

# Фильтр страницы АП ГИБДД
for row_gibdd in sheet_source_gibdd.iter_rows(min_row=2, min_col=2, values_only=True):
    list_gibdd = [cell for cell in row_gibdd]
    # Приводим формат даты рождения к ДД.ММ.ГГГГ (только если поле не пустое)
    #if list_gibdd[3] != None:
    #    list_gibdd[3] = list_gibdd[3]#.strftime("%d.%m.%Y")
    #print('АП ГИБДД:', list_gibdd)

    # Сравниваем ФИО и дату рождения между АБДЦ и Брянск
    for row_bryansk in sheet_source_bryansk.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        list_bryansk = [cell for cell in row_bryansk]
        # Приводим формат даты рождения к ДД.ММ.ГГГГ (только если поле не пустое)
        if list_bryansk[3] != None:
            list_bryansk[3] = list_bryansk[3].strftime("%d.%m.%Y")
        #print('Bryansk', list_bryansk)
        # Если строка имеет совпадения по ФИО и дате рождения, то сохраяем эту строку в итоговый файл
        if (list_gibdd[0] == list_bryansk[0] and
                list_gibdd[1] == list_bryansk[1] and
                list_gibdd[2] == list_bryansk[2] and
                list_gibdd[3] == list_bryansk[3]):
            set_gibdd.add(row_gibdd)
# Добавляем уникальные записи в новую страницу
for el in set_gibdd:
    print(f"Совпадение АП ГИБДД с Брянск: {el[0], el[1], el[2], el[3]}")
    sheet_result_gibdd.append(el)

# Фильтр страницы ЗАПРЕТНИКИ
for row_zapret in sheet_source_zapret.iter_rows(min_row=2, min_col=1, values_only=True):
    list_zapret = [cell for cell in row_zapret]
    # Приводим формат даты рождения к ДД.ММ.ГГГГ (только если поле не пустое)
    if list_zapret[3] != None:
        list_zapret[3] = list_zapret[3].strftime("%d.%m.%Y")
    #print('ЗАПРЕТНИКИ:', list_zapret)

    # Сравниваем ФИО и дату рождения между АБДЦ и Брянск
    for row_bryansk in sheet_source_bryansk.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        list_bryansk = [cell for cell in row_bryansk]
        # Приводим формат даты рождения к ДД.ММ.ГГГГ (только если поле не пустое)
        if list_bryansk[3] != None:
            list_bryansk[3] = list_bryansk[3].strftime("%d.%m.%Y")
        #print('Bryansk', list_bryansk)
        # Если строка имеет совпадения по ФИО и дате рождения, то сохраяем эту строку в итоговый файл
        if (list_zapret[0] == list_bryansk[0] and
                list_zapret[1] == list_bryansk[1] and
                list_zapret[2] == list_bryansk[2] and
                list_zapret[3] == list_bryansk[3]):
            set_zapret.add(row_zapret)
# Добавляем уникальные записи в новую страницу
for el in set_zapret:
    print(f"Совпадение ЗАПРЕТНИКИ с Брянск: {el[0], el[1], el[2], el[3]}")
    sheet_result_zapret.append(el)

# Фильтр страницы ЗАДЕРЖАНИЯ
for row_zaderj in sheet_source_zaderj.iter_rows(min_row=1, min_col=2, values_only=True):
    list_zaderj = [cell for cell in row_zaderj]
    # Приводим формат даты рождения к ДД.ММ.ГГГГ (только если поле не пустое)
    if list_zaderj[3] != None:
        list_zaderj[3] = list_zaderj[3].strftime("%d.%m.%Y")
    #print('ЗАДЕРЖАНИЯ:', list_zaderj)

    # Сравниваем ФИО и дату рождения между АБДЦ и Брянск
    for row_bryansk in sheet_source_bryansk.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        list_bryansk = [cell for cell in row_bryansk]
        # Приводим формат даты рождения к ДД.ММ.ГГГГ (только если поле не пустое)
        if list_bryansk[3] != None:
            list_bryansk[3] = list_bryansk[3].strftime("%d.%m.%Y")
        #print('Bryansk', list_bryansk)
        # Если строка имеет совпадения по ФИО и дате рождения, то сохраяем эту строку в итоговый файл
        if (list_zaderj[0] == list_bryansk[0] and
                list_zaderj[1] == list_bryansk[1] and
                list_zaderj[2] == list_bryansk[2] and
                list_zaderj[3] == list_bryansk[3]):
            set_zaderj.add(row_zaderj)
# Добавляем уникальные записи в новую страницу
for el in set_zaderj:
    print(f"Совпадение ЗАДЕРЖАНИЯ с Брянск: {el[0], el[1], el[2], el[3]}")
    sheet_result_zaderj.append(el)

# Фильтр страницы ЗАГС
for row_zags in sheet_source_zags.iter_rows(min_row=2, min_col=2, values_only=True):
    list_zags = [cell for cell in row_zags]
    # Приводим формат даты рождения к ДД.ММ.ГГГГ (только если поле не пустое)
    #if list_zags[3] != None:
    #    list_zags[3] = list_zags[3]#.strftime("%d.%m.%Y")
    #print('ЗАДЕРЖАНИЯ:', list_zags)

    # Сравниваем ФИО и дату рождения между АБДЦ и Брянск
    for row_bryansk in sheet_source_bryansk.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        list_bryansk = [cell for cell in row_bryansk]
        # Приводим формат даты рождения к ДД.ММ.ГГГГ (только если поле не пустое)
        if list_bryansk[3] != None:
            list_bryansk[3] = list_bryansk[3].strftime("%d.%m.%Y")
        #print('Bryansk', list_bryansk)
        # Если строка имеет совпадения по ФИО и дате рождения, то сохраяем эту строку в итоговый файл
        if (list_zags[0] == list_bryansk[0] and
                list_zags[1] == list_bryansk[1] and
                list_zags[2] == list_bryansk[2] and
                list_zags[3] == list_bryansk[3]):
            set_zags.add(row_zags)
# Добавляем уникальные записи в новую страницу
for el in set_zags:
    print(f"Совпадение ЗАГС с Брянск: {el[0], el[1], el[2], el[3]}")
    sheet_result_zags.append(el)


result_file_name = "result_" + source_file_name

try:
    wb_result.save(result_file_name)
except PermissionError:
    input("\nОшибка сохранения итогового файла. Возможно предыдущая версия файла уже открыта, закртойте его!\n"
          "Для выхода нажмите Enter!")
    exit()

print(f"\nИтоговый файл сохранен под названием '{result_file_name}'\n")

input("Для завершения программы нажмите Enter ")
